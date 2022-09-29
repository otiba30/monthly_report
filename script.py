# -------------------------------------
# 使い方
#
# 1. 月末書類Excelファイル元本を origin.xlsx にリネーム
# 2. シフト表ファイル作成
# 3. 下記実行
#    $ python3 ./script.py シフト表ファイル
#      -> 出力 : 【name】month月作業分_月末書類_MacVer3.xlsx
# 4. 備考欄などはExcel開いて修正
#
# 例
#
# $ python3 ./script.py shift.txt



import sys
import datetime
import openpyxl



# -------------------------------------
# 入力
#

# 参照するEXCELシート
wb = openpyxl.load_workbook('origin.xlsx')
# 勤務表シート(work schedule)
ws = wb['勤務表']
# 交通費精算シート(travel expenses)
te = wb['交通費精算']
# 参照するシフト表(shift list)
sl = sys.argv[1]

dict_sl={}
with open(sl, mode='r') as f:
    lines = f.read().splitlines()

    year, month = map(int, lines.pop(0).split())

    lines.pop(0)

    for i in lines:
        key, value = i.split()
        dict_sl[int(key)] = int(value)
f.close()



# -------------------------------------
# 変数定義
#

# 以下、個々人に合わせて書き換え

pjt_name = 'プロジェクト名'
company = '株式会社オブジェクティブコード'
work_style = '≪シフト（夜勤有）≫'
name = 'ブレイクザダークネス・オブスキュア'

# 所定開始, 所定終了, 休憩開始_1, 休憩終了_1, 休憩開始_2, 休憩終了_2,
# 開始時間, 終了時間,
# 出社or在宅,
# 備考

# 日勤(day shift)
list_ds = [[0, 30600], [0, 61200], [0, 43200], [0, 46800], None, None]\
         + [[0, 30600], [0, 61200]]\
         + ['出社']\
         + ['日勤:お仕事内容']

# 宿直(night shift)
list_ns = [[0, 59400], [1, 34200], [0, 75600], [0, 79200], [1, 10800], [1, 14400]]\
         + [[0, 59400], [1, 34200]]\
         + ['出社']\
         + ['宿直:お仕事内容']

# 日勤在宅(work remotely)
list_wr = [[0, 30600], [0, 61200], [0, 43200], [0, 46800], None, None]\
         + [[0, 30600], [0, 61200]]\
         + ['在宅']\
         + ['日勤:お仕事内容']

# 行き先, 用件, 金額, 使用機関, 移動経路_1, 移動経路_2, 往復or片道
list_te = ['OBG本社', '業務', 336, '電車', '東京', '赤坂見附', '往復']

# 書き換えここまで



# -------------------------------------
# 関数定義
#

# set work schedule
def set_ws(day, sheet):
    # 日勤・宿直・在宅で場合分け
    if dict_sl[day] == 1:
        count = 4
        list = list_ds
    elif dict_sl[day] == 2:
        count = 6
        list = list_ns
    elif dict_sl[day] == 3:
        count = 4
        list = list_wr

    # 所定開始/所定終了/休憩開始_1/休憩終了_1/休憩開始_2/休憩終了_2
    for i in range(count):
        time=datetime.timedelta(list[i][0], list[i][1])
        sheet.cell(day+7, i+3, time)

    # 開始時間/終了時間
    for i in range(2):
        time=datetime.timedelta(list[i+6][0], list[i+6][1])
        sheet.cell(day+7, i+10, time)

    # 出社or在宅
    sheet.cell(day+7, 21, list[8])
    # 備考
    sheet.cell(day+7, 24, list[9])

# set travel expenses
def set_te(day, sheet, row):
    time=datetime.date(year, month, day)
    sheet.cell(row, 2, time)

    cols = [3, 5, 6, 7, 9, 11, 12]

    for i in range(7):
        sheet.cell(row, cols[i], list_te[i])



# -------------------------------------
# 実行
#

def main():

    # ---------------------------------
    # 勤務表
    #

    ws.cell(3, 21, year)
    ws.cell(3, 23, month)
    ws.cell(4,  4, pjt_name)
    ws.cell(4, 19, company)
    ws.cell(5,  4, work_style)
    ws.cell(5, 19, name)

    for i in dict_sl:
        set_ws(i, ws)



    # ---------------------------------
    # 交通費精算
    #

    # 在宅は除外して出勤日だけ抽出
    keys = [k for k, v in dict_sl.items() if v <= 2]

    count = 0
    for i in keys:
        set_te(i, te, count+9)
        count+=1



    # ---------------------------------
    # 出力
    #

    filename = '【' + name + '】' + str(month) + '月作業分_月末書類_MacVer3.xlsx'
    wb.save(filename)
    wb.close()



if __name__ == "__main__":
    main()
